from django.db.models import Q
from django.core.cache import cache
from core.models import Question, OutOfContextQuestion, DataSource, GithubFile
from .utils import get_date_range, calculate_percentage_change, format_filter_name_for_display, map_filter_to_source
import hashlib
import json
import time
from django.core.exceptions import ValidationError
from datetime import datetime
from openpyxl import Workbook
from io import BytesIO

class AnalyticsService:
    CACHE_TTL = 15  # 15 seconds cache

    @staticmethod
    def _get_cache_key(prefix, *args):
        """Generate a cache key based on prefix and arguments."""
        key = f"analytics_{prefix}_{'_'.join(str(arg) for arg in args)}"
        return hashlib.md5(key.encode()).hexdigest()

    @staticmethod
    def get_stats_for_period(guru_type, start_date, end_date):
        """Get statistics for a specific time period with caching."""
        cache_key = AnalyticsService._get_cache_key('stats', guru_type.id, start_date.isoformat(), end_date.isoformat())
        cached_data = cache.get(cache_key)
        
        if cached_data:
            return cached_data

        questions_start = time.time()
        # Optimize queries using annotations and single database hits
        questions = Question.objects.filter(
            guru_type=guru_type,
            date_created__gte=start_date,
            date_created__lte=end_date
        ).exclude( # Exclude non slack-discord-github binge root questions, as they are duplicated from the original root question.
            ~Q(source__in=[Question.Source.SLACK.value, Question.Source.DISCORD.value, Question.Source.GITHUB.value]),
            binge_id__isnull=False,
            parent__isnull=True
        )
        
        total_questions = questions.count()
        questions_end = time.time()
        
        out_of_context_start = time.time()
        out_of_context = OutOfContextQuestion.objects.filter(
            guru_type=guru_type,
            date_created__gte=start_date,
            date_created__lte=end_date
        ).count()
        out_of_context_end = time.time()
        
        # Extract unique referenced links in a single pass
        referenced_links_start = time.time()
        referenced_links = set()
        for refs in questions.values_list('references', flat=True):
            if refs:
                for ref in refs:
                    link = ref.get('link')
                    if link:
                        referenced_links.add(link)

        # Optimize data source queries using IN clause
        referenced_sources = DataSource.objects.filter(
            guru_type=guru_type,
            url__in=referenced_links
        ).count()

        referenced_github_files = GithubFile.objects.filter(
            link__in=referenced_links
        ).count()

        referenced_links_end = time.time()
        
        result = (total_questions, out_of_context, referenced_sources + referenced_github_files)
        cache.set(cache_key, result, AnalyticsService.CACHE_TTL)
        return result

    @staticmethod
    def get_stats_data(guru_type, interval='today'):
        """Get analytics statistics with comparison to previous period."""
        cache_key = AnalyticsService._get_cache_key('stats_data', guru_type.id, interval)
        cached_data = cache.get(cache_key)
        
        if cached_data:
            return cached_data

        current_start_date, current_end_date = get_date_range(interval)
        
        current_total, current_out_of_context, current_referenced_sources = AnalyticsService.get_stats_for_period(
            guru_type, current_start_date, current_end_date
        )
        
        previous_start = current_start_date - (current_end_date - current_start_date)
        previous_total, previous_out_of_context, previous_referenced_sources = AnalyticsService.get_stats_for_period(
            guru_type, previous_start, current_start_date
        )
        
        result = {
            'total_questions': {
                'value': current_total,
                'percentage_change': calculate_percentage_change(current_total, previous_total)
            },
            'out_of_context': {
                'value': current_out_of_context,
                'percentage_change': calculate_percentage_change(current_out_of_context, previous_out_of_context)
            },
            'referenced_sources': {
                'value': current_referenced_sources,
                'percentage_change': calculate_percentage_change(current_referenced_sources, previous_referenced_sources)
            }
        }
        
        cache.set(cache_key, result, AnalyticsService.CACHE_TTL)
        return result

    @staticmethod
    def get_paginated_data(queryset, page, page_size=10):
        """Helper method to paginate queryset with optimized counting."""
        if isinstance(queryset, list):
            total_items = len(queryset)
        else:
            # Use optimized count() for querysets
            total_items = queryset.count()
        
        total_pages = (total_items + page_size - 1) // page_size
        page = min(max(1, page), total_pages) if total_pages > 0 else 1
        
        start_idx = (page - 1) * page_size
        end_idx = start_idx + page_size
        
        # Use optimized slicing for querysets
        items = queryset[start_idx:end_idx]
        
        return {
            'items': items,
            'total_pages': total_pages,
            'current_page': page,
            'total_items': total_items
        }

    @staticmethod
    def get_available_filters(metric_type):
        """Get available filters based on metric type with caching."""
        cache_key = AnalyticsService._get_cache_key('filters', metric_type)
        cached_filters = cache.get(cache_key)
        
        if cached_filters:
            return cached_filters

        if metric_type in ['questions', 'out_of_context']:
            filters = [
                {'value': 'all', 'label': 'All'},
                {'value': 'user', 'label': 'Gurubase UI'},
                {'value': 'widget', 'label': 'Widget'},
                {'value': 'api', 'label': 'API'},
                {'value': 'discord', 'label': 'Discord'},
                {'value': 'slack', 'label': 'Slack'},
                {'value': 'github', 'label': 'GitHub'},
            ]
        elif metric_type == 'referenced_sources':
            filters = [
                {'value': 'all', 'label': 'All'},
                {'value': 'github_repo', 'label': 'Codebase'},
                {'value': 'pdf', 'label': 'PDF'},
                {'value': 'website', 'label': 'Website'},
                {'value': 'youtube', 'label': 'YouTube'},
            ]
        else:
            filters = []
            
        cache.set(cache_key, filters, AnalyticsService.CACHE_TTL * 4)  # Cache filters longer as they rarely change
        return filters

    @staticmethod
    def get_data_source_questions(guru_type, data_source_url, filter_type, interval, page, search_query=None, sort_order='desc'):
        """Get questions that reference a specific data source with search and sort functionality."""
        cache_key = AnalyticsService._get_cache_key('data_source_questions', 
            guru_type.id, data_source_url, filter_type or 'all', interval, page, search_query, sort_order)
        cached_data = cache.get(cache_key)
        
        if cached_data:
            return cached_data

        start_date, end_date = get_date_range(interval)
        
        try:
            data_source = DataSource.objects.get(guru_type=guru_type, url=data_source_url)
            is_github = False
        except DataSource.DoesNotExist:
            try:
                data_source = GithubFile.objects.select_related('data_source').get(link=data_source_url)
                is_github = True
            except GithubFile.DoesNotExist:
                return None
        
        # Build the base queryset
        queryset = Question.objects.filter(
            guru_type=guru_type,
            date_created__gte=start_date,
            date_created__lte=end_date,
            references__contains=[{'link': data_source_url}]
        ).exclude( # Exclude non slack-discord-github binge root questions, as they are duplicated from the original root question.
            ~Q(source__in=[Question.Source.SLACK.value, Question.Source.DISCORD.value, Question.Source.GITHUB.value]),
            binge_id__isnull=False,
            parent__isnull=True
        )
        
        # Apply filter type if specified
        if filter_type and filter_type != 'all':
            source_value = map_filter_to_source(filter_type)
            if source_value:
                queryset = queryset.filter(source__iexact=source_value)
        
        # Apply search filter if query exists
        if search_query:
            queryset = queryset.filter(user_question__icontains=search_query)
        
        # Apply sorting
        order_by = 'date_created' if sort_order == 'asc' else '-date_created'
        queryset = queryset.order_by(order_by)
        
        paginated_data = AnalyticsService.get_paginated_data(queryset, page)
        
        results = [{
            'date': item.date_created.isoformat(),
            'title': item.user_question,
            'truncated_title': item.user_question[:75] + '...' if len(item.user_question) > 75 else item.user_question,
            'link': item.frontend_url,
            'source': format_filter_name_for_display(item.source)
        } for item in paginated_data['items']]
        
        result = {
            'results': results,
            'total_pages': paginated_data['total_pages'],
            'current_page': paginated_data['current_page'],
            'total_items': paginated_data['total_items'],
            'available_filters': AnalyticsService.get_available_filters('questions')
        }
        
        cache.set(cache_key, result, AnalyticsService.CACHE_TTL)
        return result

    @staticmethod
    def _get_filtered_questions(guru_type, start_date, end_date, filter_type=None, search_query=None, sort_order='desc'):
        """Get filtered questions based on criteria."""
        queryset = Question.objects.filter(
            guru_type=guru_type,
            date_created__gte=start_date,
            date_created__lte=end_date
        ).exclude(
            ~Q(source__in=[Question.Source.SLACK.value, Question.Source.DISCORD.value, Question.Source.GITHUB.value]),
            binge_id__isnull=False,
            parent__isnull=True
        )

        if filter_type and filter_type != 'all':
            source_value = map_filter_to_source(filter_type)
            if source_value:
                queryset = queryset.filter(source__iexact=source_value)

        if search_query:
            queryset = queryset.filter(user_question__icontains=search_query)

        order_by = 'date_created' if sort_order == 'asc' else '-date_created'
        return queryset.order_by(order_by)

    @staticmethod
    def _get_filtered_out_of_context(guru_type, start_date, end_date, filter_type=None, search_query=None, sort_order='desc'):
        """Get filtered out of context questions based on criteria."""
        queryset = OutOfContextQuestion.objects.filter(
            guru_type=guru_type,
            date_created__gte=start_date,
            date_created__lte=end_date
        )

        if filter_type and filter_type != 'all':
            source_value = map_filter_to_source(filter_type)
            if source_value:
                queryset = queryset.filter(source__iexact=source_value)

        if search_query:
            queryset = queryset.filter(user_question__icontains=search_query)

        order_by = 'date_created' if sort_order == 'asc' else '-date_created'
        return queryset.order_by(order_by)

    @staticmethod
    def _get_filtered_referenced_sources(guru_type, start_date, end_date, filter_type=None, search_query=None, sort_order='desc'):
        """Get filtered referenced sources based on criteria."""
        # Get questions with references
        questions = Question.objects.filter(
            guru_type=guru_type,
            date_created__gte=start_date,
            date_created__lte=end_date
        ).exclude(
            ~Q(source__in=[Question.Source.SLACK.value, Question.Source.DISCORD.value, Question.Source.GITHUB.value]),
            binge_id__isnull=False,
            parent__isnull=True
        ).values('references')

        # Extract referenced links and count occurrences
        reference_counts = {}
        referenced_links = set()

        for question in questions:
            for ref in question.get('references', []):
                link = ref.get('link')
                if link:
                    referenced_links.add(link)
                    reference_counts[link] = reference_counts.get(link, 0) + 1

        # Get data sources based on filter
        if filter_type == 'all' or not filter_type:
            data_sources = DataSource.objects.filter(
                guru_type=guru_type,
                url__in=referenced_links,
                status=DataSource.Status.SUCCESS
            )
            github_files = GithubFile.objects.filter(
                link__in=referenced_links,
                data_source__status=DataSource.Status.SUCCESS
            ).select_related('data_source')
        elif filter_type == 'github_repo':
            data_sources = []
            github_files = GithubFile.objects.filter(
                link__in=referenced_links,
                data_source__status=DataSource.Status.SUCCESS
            ).select_related('data_source')
        else:
            data_sources = DataSource.objects.filter(
                guru_type=guru_type,
                url__in=referenced_links,
                type__iexact=filter_type,
                status=DataSource.Status.SUCCESS
            )
            github_files = []

        # Combine and process results
        combined_sources = []

        for ds in data_sources:
            if ds.type and ds.title and ds.url:
                combined_sources.append({
                    'date': ds.date_created.isoformat(),
                    'type': format_filter_name_for_display(ds.type),
                    'title': ds.title,
                    'url': ds.url,
                    'reference_count': reference_counts.get(ds.url, 0)
                })

        for gf in github_files:
            if gf.data_source and gf.data_source.date_created and gf.title and gf.link:
                combined_sources.append({
                    'date': gf.data_source.date_created.isoformat(),
                    'type': 'Codebase',
                    'title': gf.title,
                    'url': gf.link,
                    'reference_count': reference_counts.get(gf.link, 0)
                })

        # Filter out duplicate links, keeping the one with highest reference count
        seen_links = {}
        unique_sources = []
        for source in combined_sources:
            link = source['url']
            if link not in seen_links:
                seen_links[link] = source
                unique_sources.append(source)
            else:
                existing = seen_links[link]
                if source['reference_count'] > existing['reference_count']:
                    unique_sources.remove(existing)
                    seen_links[link] = source
                    unique_sources.append(source)

        # Apply search filter if needed
        if search_query:
            unique_sources = [
                source for source in unique_sources 
                if search_query.lower() in source['title'].lower()
            ]

        # Sort results
        if sort_order == 'asc':
            unique_sources.sort(key=lambda x: (x['reference_count'], x['date']))
        else:
            unique_sources.sort(key=lambda x: (x['reference_count'], x['date']), reverse=True)

        return unique_sources

    @staticmethod
    def _prepare_xlsx_data(results):
        """Prepare data for Excel export with proper headers and formatting."""
        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)
        
        # Questions sheet
        if results['questions']:
            ws = wb.create_sheet('Questions')
            headers = ['Datetime', 'Source', 'Question', 'Trust Score', 'Follow-up']
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = cell.font.copy(bold=True)
            
            # Write data
            for row_idx, question in enumerate(results['questions'], 2):
                ws.cell(row=row_idx, column=1, value=question.date_created.strftime('%Y-%m-%d %H:%M'))
                ws.cell(row=row_idx, column=2, value=format_filter_name_for_display(question.source))
                
                # Question with hyperlink
                cell = ws.cell(row=row_idx, column=3, value=question.user_question)
                cell.hyperlink = question.frontend_url
                cell.style = 'Hyperlink'
                
                # Trust score
                ws.cell(row=row_idx, column=4, value=f'{question.trust_score:.2f}' if question.trust_score is not None else '')
                ws.cell(row=row_idx, column=5, value=question.parent is not None)
        
        # Unable to Answer sheet
        if results['unable_to_answer']:
            ws = wb.create_sheet('Unable to Answer')
            headers = ['Datetime', 'Source', 'Question']
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = cell.font.copy(bold=True)
            
            # Write data
            for row_idx, ooc in enumerate(results['unable_to_answer'], 2):
                ws.cell(row=row_idx, column=1, value=ooc.date_created.strftime('%Y-%m-%d %H:%M'))
                ws.cell(row=row_idx, column=2, value=format_filter_name_for_display(ooc.source))
                ws.cell(row=row_idx, column=3, value=ooc.user_question)
        
        # References sheet
        if results['references']:
            ws = wb.create_sheet('References')
            headers = ['Last Update Date', 'Data Source Type', 'Data Source Title', 'Referenced Count']
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = cell.font.copy(bold=True)
            
            # Write data
            for row_idx, source in enumerate(results['references'], 2):
                ws.cell(row=row_idx, column=1, value=datetime.fromisoformat(source['date']).strftime('%Y-%m-%d %H:%M'))
                
                ws.cell(row=row_idx, column=2, value=source['type'])

                # Data source title with hyperlink
                cell = ws.cell(row=row_idx, column=3, value=source['title'])
                cell.hyperlink = source['url']
                cell.style = 'Hyperlink'
                
                ws.cell(row=row_idx, column=4, value=source['reference_count'])
        
        # Auto-adjust column widths for all sheets
        for ws in wb.worksheets:
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = min(adjusted_width, 50)  # Cap width at 50
        
        # Save to bytes buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()

    @staticmethod
    def _prepare_csv_data(results):
        """Prepare data for CSV export with proper headers and formatting.
        Creates separate CSV files for each metric type and returns them as a zip file."""
        import csv
        from io import StringIO, BytesIO
        import zipfile

        zip_buffer = BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Questions section
            if results['questions']:
                output = StringIO()
                writer = csv.writer(output)
                writer.writerow(['Datetime', 'Source', 'Question', 'Trust Score', 'Follow-up', 'URL'])
                
                for question in results['questions']:
                    if len(question.user_question) > 75:
                        x = 2
                    writer.writerow([
                        question.date_created.strftime('%Y-%m-%d %H:%M'),
                        format_filter_name_for_display(question.source),
                        question.user_question,
                        f'{question.trust_score:.2f}' if question.trust_score is not None else '',
                        'Yes' if question.parent else 'No',
                        question.frontend_url
                    ])
                
                zip_file.writestr('questions.csv', output.getvalue())

            # Unable to Answer section
            if results['unable_to_answer']:
                output = StringIO()
                writer = csv.writer(output)
                writer.writerow(['Datetime', 'Source', 'Question'])
                
                for ooc in results['unable_to_answer']:
                    writer.writerow([
                        ooc.date_created.strftime('%Y-%m-%d %H:%M'),
                        format_filter_name_for_display(ooc.source),
                        ooc.user_question
                    ])
                
                zip_file.writestr('unable_to_answer.csv', output.getvalue())

            # References section
            if results['references']:
                output = StringIO()
                writer = csv.writer(output)
                writer.writerow(['Last Update Date', 'Data Source Type', 'Data Source Title', 'URL', 'Referenced Count'])
                
                for source in results['references']:
                    writer.writerow([
                        datetime.fromisoformat(source['date']).strftime('%Y-%m-%d %H:%M'),
                        source['type'],
                        source['title'],
                        source['url'],
                        source['reference_count']
                    ])
                
                zip_file.writestr('references.csv', output.getvalue())

        return zip_buffer.getvalue()

    @staticmethod
    def _prepare_json_data(results):
        """Prepare data for JSON export with proper formatting."""
        export_data = {
            'questions': [{
                'datetime': question.date_created.strftime('%Y-%m-%d %H:%M'),
                'source': format_filter_name_for_display(question.source),
                'question': question.user_question,
                'trust_score': float(f'{question.trust_score:.2f}') if question.trust_score is not None else None,
                'follow_up': question.parent is not None,
                'url': question.frontend_url
            } for question in results['questions']],
            
            'unable_to_answer': [{
                'datetime': ooc.date_created.strftime('%Y-%m-%d %H:%M'),
                'source': format_filter_name_for_display(ooc.source),
                'question': ooc.user_question
            } for ooc in results['unable_to_answer']],
            
            'references': [{
                'last_update_date': datetime.fromisoformat(source['date']).strftime('%Y-%m-%d %H:%M'),
                'data_source_type': source['type'],
                'data_source_title': source['title'],
                'referenced_count': source['reference_count'],
                'url': source['url']
            } for source in results['references']]
        }
        
        return json.dumps(export_data, indent=2).encode('utf-8')

    @staticmethod
    def fetch_export_data(guru_type, interval, filters):
        """Fetch data for export without any formatting."""
        start_date, end_date = get_date_range(interval)
        results = {
            'questions': [],
            'unable_to_answer': [],
            'references': []
        }

        # Handle questions export
        if 'questions' in filters:
            questions_filter = filters['questions']
            queryset = AnalyticsService._get_filtered_questions(
                guru_type, start_date, end_date, questions_filter
            )
            results['questions'] = list(queryset)

        # Handle out of context questions export
        if 'out_of_context' in filters:
            ooc_filter = filters['out_of_context']
            queryset = AnalyticsService._get_filtered_out_of_context(
                guru_type, start_date, end_date, ooc_filter
            )
            results['unable_to_answer'] = list(queryset)

        # Handle referenced sources export
        if 'referenced_sources' in filters:
            sources_filter = filters['referenced_sources']
            results['references'] = AnalyticsService._get_filtered_referenced_sources(
                guru_type, start_date, end_date, sources_filter
            )

        # Sort results by date
        if results['questions']:
            results['questions'].sort(key=lambda x: x.date_created, reverse=True)
        if results['unable_to_answer']:
            results['unable_to_answer'].sort(key=lambda x: x.date_created, reverse=True)
        if results['references']:
            results['references'].sort(key=lambda x: x['reference_count'], reverse=True)

        return results

    @staticmethod
    def export_analytics_data(guru_type, export_type, interval, filters):
        """Export analytics data in the specified format."""
        if export_type not in ['xlsx', 'csv', 'json']:
            raise ValidationError('Export type must be one of: xlsx, csv, json')

        # Fetch the data
        results = AnalyticsService.fetch_export_data(guru_type, interval, filters)
        
        # Format and return data based on export type
        if export_type == 'xlsx':
            return AnalyticsService._prepare_xlsx_data(results)
        elif export_type == 'csv':
            return AnalyticsService._prepare_csv_data(results)
        else:  # json
            return AnalyticsService._prepare_json_data(results) 